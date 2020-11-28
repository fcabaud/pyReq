#!/usr/bin/python
# Goal : to extract dictionnary in an xlsx file from a list of requirements
from pyReq import *
from openpyxl import Workbook
import argparse

C_SEPARATOR = ','

class ReqGetRedmine(pyReq):
  """ class for export of requirements in excel 
     :arg pyReq: json input requirements file 
     :type pyReq: string
  """
  def get_redmine_csv(self, listOfReq = [], redmineCsvName = 'getRedmine.csv'):
    """ Get redmine csv file
    
       :arg redmineCsvName: csv file containing requirements and attributes 
         with C_SEPARATOR as separator 
       :type redmineCsvName: string
    """
    #dictReq = self.get(listOfReq)
    #print(listOfReq)
    if listOfReq == []:
      listOfReq = self.reqDict.keys()
    print(listOfReq)
    #print("Request list  %s",";".join(set(listOfReq)))
    self.raiseOnNotFoundRequirement(listOfReq)    
    first_line = "Tracker,Status,Subject,Description\n"

    fp = open(redmineCsvName, "w")
    fp.write(first_line)
    # Get all attributes name for all tags
    """attributesDict={}
    nb=0
    for tag in listOfReq:
      for attribute in self.reqDict[tag][C_KEY_ATTRIBUTES].keys():
        if attribute not in attributesDict:
          attributesDict[attribute] = nb
          # in line 0, write attribute name
          ws.cell(row = 1, column = 5 + nb).value = attribute
          nb = nb + 1 """

    # Iterate over the data and write it out row by row.
    #print(self.reqDict.keys())
    #
    # 122121, User Story, New, Subject, Description
    for tag in listOfReq:
        #print(self.reqDict[tag])
        cvs_line = 'User Story' + C_SEPARATOR
        cvs_line = cvs_line + 'New' + C_SEPARATOR

        if C_SEPARATOR in tag:
            print("Separator %s is found in TAG %s, so it can not be provided in csv file"%(C_SEPARATOR, tag))
        if '"' in tag:
            print(' Character " is found in TAG %s, so it can not be provided in csv file as subject field for redmine import because syntax is "requiment"'%(tag))
        cvs_line = cvs_line + '"' + tag + '"' + C_SEPARATOR

        if self.reqDict[tag][C_KEY_BODY]: # not None
            if C_SEPARATOR in self.reqDict[tag][C_KEY_BODY]:
                print("Separator %s is found in requirement %s of TAG %s, so it can not be provided in csv file"%(C_SEPARATOR, self.reqDict[tag][C_KEY_BODY], tag))
            if '"' in self.reqDict[tag][C_KEY_BODY]:
                print(' Character " is found in requirement %s of TAG %s, so it can not be provided in csv file as description field for redmine import because syntax is "description"'%(self.reqDict[tag][C_KEY_BODY], tag))
            cvs_line = cvs_line + '"' + self.reqDict[tag][C_KEY_BODY] + '"'
        else:
            cvs_line = cvs_line + '" "'

        cvs_line = cvs_line + "\n"
        fp.write(cvs_line)
        # self.reqDict[tag][C_KEY_BODY]
        # self.reqDict[tag][C_KEY_DOCUMENT]
        # self.reqDict[tag][C_KEY_COVERAGE]
        #ws.cell(row = line, column = 5).value = ";".join(self.reqDict[tag][C_KEY_ATTRIBUTES])
        #for attribute in self.reqDict[tag][C_KEY_ATTRIBUTES].keys():
        #  ws.cell(row = line, column = 5+attributesDict[attribute]).value = self.reqDict[tag][C_KEY_ATTRIBUTES][attribute]

    fp.close()
    print("Write redmine csv file %s"%redmineCsvName)

def test():
  getReqInstance = ReqGetRedmine(C_PATH_WORK+"docExample.json")
  #listOfTags = ['RQT_0001','RQT_0003']
  #print(getReqInstance.getKeys())
  # Example 1 : get all requirements not covered of sprint 2 : what validation team has to do
  listOfTagsSprint2 = getReqInstance.getListReqFromAttribute("attributeSprint", 2)
  for tag in listOfTagsSprint2:
    if getReqInstance[tag][C_KEY_COVERAGE] == []:
      print("%s Not covered"%tag)  
  getReqInstance.get_redmine_csv(listOfTagsSprint2, C_PATH_OUT+'reqListSprint2NotCovered.xlsx')  
  # Example 2 : get all requirements covered by a KO test : what development team has to do
  listOfTagsSprint1 = getReqInstance.getListReqFromAttribute("attributeStatus", "KO")
  getReqInstance.get_redmine_csv(listOfTagsSprint1, C_PATH_OUT+'reqListStatusKO.xlsx')  
    
if __name__ == '__main__':
  #test()
  parser = argparse.ArgumentParser(description='json2redmine %sdocExample.json %s\exampleRedmine.csv'%(C_PATH_WORK, C_PATH_OUT))
  parser.add_argument('jsonFileInput', action="store")
  parser.add_argument('redmineCsvOutput', action="store")
  result = parser.parse_args()
  arguments = dict(result._get_kwargs())  
  #print(arguments['xlsxFileInput'])
  #print(arguments['jsonFileOutput'])
  getReqInstance = ReqGetRedmine(arguments['jsonFileInput'])
  listOfTagsSprint1 = getReqInstance.getListReqFromAttribute("attributeStatus", "KO")
  #getReqInstance.get_redmine_csv(listOfTagsSprint1, arguments['redmineCsvOutput'])  
  # [] means all requirements
  getReqInstance.get_redmine_csv([], arguments['redmineCsvOutput'])  
  del(getReqInstance)
