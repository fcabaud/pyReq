#!/usr/bin/python
# Goal : to extract dictionnary in an xlsx file from a list of requirements
from pyReq import *
from openpyxl import Workbook
import argparse

class ReqGetXlsx(pyReq):
  """ class for export of requirements in excel 
     :arg pyReq: json input requirements file 
     :type pyReq: string
  """
  def get_xlsx(self, listOfReq = [], excelName = 'getReq.xlsx'):
    """ Get xlslx file 
    
       :arg excelName: xlsx file containing requirements and attributes
       :type excelName: string
    """
    #dictReq = self.get(listOfReq)
    #print(listOfReq)
    if listOfReq == []:
      listOfReq = self.reqDict.keys()
    print(listOfReq)
    #print("Request list  %s",";".join(set(listOfReq)))
    self.raiseOnNotFoundRequirement(listOfReq)    
    wb = Workbook()
    ws = wb.get_active_sheet()
    ws.title = "Requirements"

    #  from the first cell. Rows and columns are zero indexed.
    line = 1
    col = 0

    ws.cell(row = 0, column = 0).value = C_KEY_TAG
    ws.cell(row = 0, column = 1).value = C_KEY_BODY
    ws.cell(row = 0, column = 2).value = C_KEY_DOCUMENT
    ws.cell(row = 0, column = 3).value = C_KEY_COVERAGE
    #ws.cell(row = 0, column = 4).value = C_KEY_ATTRIBUTES
    # Get all attributes name for all tags
    attributesDict={}
    nb=0
    for tag in listOfReq:
      for attribute in self.reqDict[tag][C_KEY_ATTRIBUTES].keys():
        if attribute not in attributesDict:
          attributesDict[attribute] = nb
          # in line 0, write attribute name
          ws.cell(row = 0, column = 4 + nb).value = attribute
          nb = nb + 1

    # Iterate over the data and write it out row by row.
    #print(self.reqDict.keys())
    for tag in listOfReq:
      #print(self.reqDict[tag])
      ws.cell(row = line, column = 0).value = tag
      ws.cell(row = line, column = 1).value = self.reqDict[tag][C_KEY_BODY]
      ws.cell(row = line, column = 2).value = self.reqDict[tag][C_KEY_DOCUMENT]
      ws.cell(row = line, column = 3).value = ";".join(self.reqDict[tag][C_KEY_COVERAGE])
      #ws.cell(row = line, column = 4).value = ";".join(self.reqDict[tag][C_KEY_ATTRIBUTES])
      for attribute in self.reqDict[tag][C_KEY_ATTRIBUTES].keys():
        ws.cell(row = line, column = 4+attributesDict[attribute]).value = self.reqDict[tag][C_KEY_ATTRIBUTES][attribute]
      line += 1

    wb.save(excelName)
    print("Write xlsx file %s"%excelName)

def test():
  getReqInstance = ReqGetXlsx(C_PATH_WORK+"docExample.json")
  #listOfTags = ['RQT_0001','RQT_0003']
  #print(getReqInstance.getKeys())
  # Example 1 : get all requirements not covered of sprint 2 : what validation team has to do
  listOfTagsSprint2 = getReqInstance.getListReqFromAttribute("attributeSprint", 2)
  for tag in listOfTagsSprint2:
    if getReqInstance[tag][C_KEY_COVERAGE] == []:
      print("%s Not covered"%tag)  
  getReqInstance.get_xlsx(listOfTagsSprint2, C_PATH_OUT+'reqListSprint2NotCovered.xlsx')  
  # Example 2 : get all requirements covered by a KO test : what development team has to do
  listOfTagsSprint1 = getReqInstance.getListReqFromAttribute("attributeStatus", "KO")
  getReqInstance.get_xlsx(listOfTagsSprint1, C_PATH_OUT+'reqListStatusKO.xlsx')  
    
if __name__ == '__main__':
  #test()
  parser = argparse.ArgumentParser(description='json2xlsx %sdocExample.json %s\reqListStatusKO.xlsx'%(C_PATH_WORK, C_PATH_OUT))
  parser.add_argument('jsonFileInput', action="store")
  parser.add_argument('xlsxFileOutput', action="store")
  result = parser.parse_args()
  arguments = dict(result._get_kwargs())  
  #print(arguments['xlsxFileInput'])
  #print(arguments['jsonFileOutput'])
  getReqInstance = ReqGetXlsx(arguments['jsonFileInput'])
  listOfTagsSprint1 = getReqInstance.getListReqFromAttribute("attributeStatus", "KO")
  getReqInstance.get_xlsx(listOfTagsSprint1, arguments['xlsxFileOutput'])  
