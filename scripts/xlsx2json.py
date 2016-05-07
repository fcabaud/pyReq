#!/usr/bin/python
# Goal : Add to a list of existing requirements: 
#               a coverage and attributes 
import os
from pyReq import *
from openpyxl import load_workbook
import argparse

class ReqCoverageAttributes(pyReq):
  """ class for management of requirements coverages and attributes 
     :arg json file: input and output requirements file 
     :type json filen: string
  """
  def set_xlsx(self, excelName = 'setReq.xlsx'):
    """ provides xlsx file excelName containing coverage and attributes
    
       :arg excelName: xlsx file containing coverage and attributes
       :type excelName: string
    """
    print(excelName)
    wb = load_workbook(excelName)
    #ws = wb.get_active_sheet()
    ws = wb.active
    # 1) Get the column number from the name and check mandatories name
    line1Names = {}
    #for item in range(ws.get_highest_column()):
    for item in range(ws.max_column): 
      if ws.cell(row = 1, column = item).value != None:
        #print(ws.cell(row = 0, column = item).value)
        line1Names[ws.cell(row = 1, column = item).value] = item
    #print(line1Names)
    # Check mandatories fields
    self.check_mandatories_field_names(excelName, line1Names)    
    # Get attributes names
    attributesListName=[]
    #print(ws.get_highest_column())
    for item in line1Names.keys():
      #print(item)
      if C_KEY_ATTRIBUTES in item:
        attributesListName.append(item)
    #print(attributesListName)
    # 2) Import data
    for line in range(ws.max_row):
      if line > 1:
        # 2.1) Get tag field
        tag      = ws.cell(row = line, column = line1Names[C_KEY_TAG]).value
        # 2.2) Get body
        body     = ws.cell(row = line, column = line1Names[C_KEY_BODY]).value
        # 2.2) Get document name
        document = ws.cell(row = line, column = line1Names[C_KEY_DOCUMENT]).value
        # 3) Optionnal fields
        coverage   = []
        if C_KEY_COVERAGE in line1Names:
          if ws.cell(row = line, column = line1Names[C_KEY_COVERAGE]).value != None:
            coverage   = ws.cell(row = line, column = line1Names[C_KEY_COVERAGE]).value.split(";")
        #print("coverage")
        #print(self.reqDict[tag][C_KEY_COVERAGE])
        #self.reqDict[tag][C_KEY_ATTRIBUTES] = ws.cell(row = line, column = line1Names[C_KEY_ATTRIBUTES]).value
        attributesDict = {}
        for attribute in attributesListName:
          if ws.cell(row = line, column = line1Names[attribute]).value != None:
            #self.reqDict[tag][C_KEY_ATTRIBUTES][attribute]   = ws.cell(row = line, column = line1Names[attribute]).value
            attributesDict[attribute]   = ws.cell(row = line, column = line1Names[attribute]).value

        # write in internal structure
        self.add(tag, document, body, coverage, attributesDict)
  #
  def check_mandatories_field_names(self, excelName, line1Names):
    """ Check 3 fields are mandatories : tag, body and document name
      an error is raised if one field is missing 
    """
    if C_KEY_TAG not in line1Names: 
      err = " '%s' not found in first line of %s"%(C_KEY_TAG, excelName)
      raise error(err)
    if C_KEY_BODY not in line1Names: 
      err = " '%s' not found in first line of %s"%(C_KEY_BODY, excelName)
      raise error(err)
    if C_KEY_DOCUMENT not in line1Names: 
      err = " '%s' not found in first line of %s"%(C_KEY_DOCUMENT, excelName)
      raise error(err)

def test():        
  getReqInstance = ReqCoverageAttributes(C_PATH_WORK+"docExample.json")
  getReqInstance.set_xlsx(C_PATH_IN+'reqListCoverage.xlsx')  
  getReqInstance.set_xlsx(C_PATH_IN+'reqListSprints.xlsx')  
  getReqInstance.set_xlsx(C_PATH_IN+'reqListSprints_OnlyStatus.xlsx')  
        
if __name__ == '__main__':
  #test()
  parser = argparse.ArgumentParser(description='xlsx2json %sreqListCoverage.xlsx %sdocExample.json'%(C_PATH_IN, C_PATH_WORK))
  parser.add_argument('xlsxFileInput', action="store")
  parser.add_argument('jsonFile', action="store")
  result = parser.parse_args()
  arguments = dict(result._get_kwargs())  
  #print(arguments['xlsxFileInput'])
  #print(arguments['jsonFile'])
  getReqInstance = ReqCoverageAttributes(arguments['jsonFile'])
  getReqInstance.set_xlsx(arguments['xlsxFileInput'])  
  
